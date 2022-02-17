import datetime
import openpyxl


book = openpyxl.load_workbook("task_support.xlsx")
sheet = book.worksheets[1]
answers_dict = {}


# 1
num1 = [sheet[i][1].value for i in range(3, sheet.max_row+1)]
counter_num1 = 0
for num in num1:
    if num % 2 == 0:
        counter_num1 += 1

answers_dict.update({sheet[2][1].value: counter_num1})


# 2
num2 = [sheet[i][2].value for i in range(3, sheet.max_row+1)]
counter_num2 = 0


def is_prime(n):
    x = 2
    while n % x != 0:
        x += 1
    return x == n


for num in num2:
    if is_prime(num):
        counter_num2 += 1

answers_dict.update({sheet[2][2].value: counter_num2})


# 3
num3 = [float(sheet[i][3].value.replace(' ', '').replace(',', '.')) for i in range(3, sheet.max_row+1)]
counter_num3 = 0
for num in num3:
    if num < 0.5:
        counter_num3 += 1

answers_dict.update({sheet[2][3].value: counter_num3})


# 4
num4 = [str(sheet[i][4].value.strip()) for i in range(3, sheet.max_row+1)]
counter_num4 = 0

for num in num4:
    if num.startswith('Tue'):
        counter_num4 += 1

answers_dict.update({sheet[2][4].value: counter_num4})


# 5
num5 = [str(sheet[i][5].value.strip()) for i in range(3, sheet.max_row+1)]
counter_num5 = 0
for num in num5:
    d = datetime.datetime.strptime(num[:10], '%Y-%m-%d')
    if d.isoweekday() == 2:
        counter_num5 += 1
answers_dict.update({sheet[2][5].value: counter_num5})


# 6
num6 = [str(sheet[i][6].value.strip()) for i in range(3, sheet.max_row+1)]
counter_num6 = 0
for num in num6:
    if int(num[3:5]) > 23:
        d = datetime.datetime.strptime(num, '%m-%d-%Y')
        if d.isoweekday() == 2:
            counter_num6 += 1

answers_dict.update({sheet[2][6].value: counter_num6})


if 'Answers' not in book:
    book.create_sheet('Answers')

answers = book['Answers']

for i in range(1, 7):
    answers[f'A{i}'] = list(answers_dict.keys())[i-1]
    answers[f'B{i}'] = list(answers_dict.values())[i-1]

book.save('task_support.xlsx')
